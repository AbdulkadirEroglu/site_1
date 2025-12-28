import logging
from datetime import datetime
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple, Union
from uuid import uuid4
import io
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from markupsafe import Markup
import markdown as md

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.csrf import ensure_csrf_token, validate_csrf_token
from app.core.rate_limit import RateLimiter
from app.core.security import verify_password
from app.core.text import render_rich_text
from app.db.models import AdminUser, Category, Lead, Product, ProductImage, SiteMetric
from app.db.session import get_db

templates = Jinja2Templates(directory="app/templates")
templates.env.filters["markdown"] = lambda text: Markup(md.markdown(text or "", extensions=["extra", "sane_lists"]))
templates.env.filters["richtext"] = render_rich_text
templates.env.globals["static_version"] = settings.static_version

router = APIRouter(prefix="/admin", tags=["Admin"])

logger = logging.getLogger("app.admin")
settings = get_settings()
login_rate_limiter = RateLimiter(
    max_requests=settings.login_rate_limit_max_attempts,
    window_seconds=settings.login_rate_limit_window_seconds,
)

UPLOAD_ROOT = Path("app/static/uploads/products")
UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}


def _get_admin_from_session(request: Request, db: Session) -> Optional[AdminUser]:
    admin_id = request.session.get("admin_user_id")
    if not admin_id:
        return None

    admin = db.get(AdminUser, admin_id)
    if not admin or not admin.is_active:
        request.session.pop("admin_user_id", None)
        return None
    return admin


def _ensure_admin(request: Request, db: Session) -> Optional[AdminUser]:
    return _get_admin_from_session(request, db)


def _slugify(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9\s-]", "", value)
    value = re.sub(r"[\s_-]+", "-", value)
    return value.strip("-")


class ImageUploadError(Exception):
    """Raised when an uploaded image cannot be processed."""


async def _process_new_uploads(
    files: Optional[List[UploadFile]],
    *,
    default_alt: str,
    starting_order: int = 0,
) -> List[ProductImage]:
    saved_images: List[ProductImage] = []
    if not files:
        return saved_images

    order = starting_order
    for upload in files:
        if not upload or not upload.filename:
            continue

        suffix = Path(upload.filename).suffix.lower()
        if suffix not in ALLOWED_IMAGE_EXTENSIONS:
            await upload.close()
            raise ImageUploadError("Unsupported image type. Upload JPG, PNG, GIF, or WebP files.")

        data = await upload.read()
        await upload.close()
        if not data:
            continue

        filename = f"{uuid4().hex}{suffix}"
        destination = UPLOAD_ROOT / filename
        destination.write_bytes(data)

        saved_images.append(
            ProductImage(
                image_url=f"/static/uploads/products/{filename}",
                alt_text=default_alt,
                sort_order=order,
            )
        )
        order += 1

    return saved_images


def _remove_image_file(image_url: Optional[str]) -> None:
    if not image_url:
        return
    prefix = "/static/"
    if image_url.startswith(prefix):
        relative_path = image_url[len(prefix) :]
    else:
        relative_path = image_url.lstrip("/")

    full_path = Path("app/static") / relative_path
    try:
        full_path.resolve().relative_to(Path("app/static").resolve())
    except ValueError:
        return

    if full_path.exists():
        full_path.unlink()


def _build_context(request: Request, extra: Dict[str, Any]) -> Dict[str, Any]:
    context = {"request": request, "csrf_token": ensure_csrf_token(request)}
    context.update(extra)
    return context


def _client_identifier(request: Request) -> str:
    if request.client and request.client.host:
        return request.client.host
    return "unknown"


def _coerce_uploads(
    files: Union[UploadFile, List[UploadFile], None],
) -> List[UploadFile]:
    if not files:
        return []
    if isinstance(files, list):
        return [file for file in files if file]
    return [files]


def _category_parent_options(
    db: Session,
    *,
    include_predicate: Optional[Callable[[Category], bool]] = None,
) -> List[Dict[str, str]]:
    categories = db.scalars(select(Category).order_by(Category.order.asc(), Category.name.asc())).all()
    children_map: Dict[Optional[int], List[Category]] = defaultdict(list)
    for category in categories:
        children_map[category.parent_id].append(category)

    for siblings in children_map.values():
        siblings.sort(key=lambda c: (c.order, c.name.lower()))

    options: List[Dict[str, str]] = []
    visited: Set[int] = set()

    def _visit(node: Category, depth: int) -> None:
        prefix = "-- " * depth
        if not include_predicate or include_predicate(node):
            label = f"{prefix}{node.name}" if prefix else node.name
            options.append({"id": node.id, "id_str": str(node.id), "label": label})
        visited.add(node.id)
        for child in children_map.get(node.id, []):
            _visit(child, depth + 1)

    for root in children_map.get(None, []):
        _visit(root, 0)

    for category in categories:
        if category.id not in visited:
            _visit(category, 0)

    return options


def _category_tree_with_stats(db: Session) -> List[Dict[str, Any]]:
    rows = db.execute(
        select(
            Category.id,
            Category.name,
            Category.slug,
            Category.is_active,
            Category.parent_id,
            Category.order,
            Category.level,
            func.count(Product.id).label("product_count"),
        )
        .outerjoin(Product, Product.category_id == Category.id)
        .group_by(
            Category.id,
            Category.name,
            Category.slug,
            Category.is_active,
            Category.parent_id,
            Category.order,
            Category.level,
        )
    ).all()

    nodes: Dict[int, Dict[str, Any]] = {}
    children_map: Dict[Optional[int], List[Dict[str, Any]]] = defaultdict(list)

    for row in rows:
        node = {
            "id": row.id,
            "name": row.name,
            "slug": row.slug,
            "is_active": row.is_active,
            "parent_id": row.parent_id,
            "order": row.order,
            "level": row.level,
            "product_count": row.product_count,
        }
        nodes[row.id] = node
        children_map[row.parent_id].append(node)

    for siblings in children_map.values():
        siblings.sort(key=lambda item: (item["order"], item["name"].lower()))

    ordered: List[Dict[str, Any]] = []
    visited: Set[int] = set()

    def _visit(node: Dict[str, Any], depth: int) -> None:
        node["depth"] = depth
        ordered.append(node)
        visited.add(node["id"])
        for child in children_map.get(node["id"], []):
            _visit(child, depth + 1)

    for root in children_map.get(None, []):
        _visit(root, 0)

    for node in nodes.values():
        if node["id"] not in visited:
            _visit(node, 0)

    return ordered


def _get_metric(db: Session, key: str) -> int:
    metric = db.scalar(select(SiteMetric).where(SiteMetric.key == key))
    return metric.value if metric else 0


def _render_categories_page(
    request: Request,
    admin: AdminUser,
    db: Session,
    *,
    search_term: str = "",
    message: Optional[str] = None,
    message_kind: str = "info",
) -> HTMLResponse:
    categories = _category_tree_with_stats(db)
    trimmed = search_term.strip()
    if trimmed:
        lowered = trimmed.lower()
        categories = [
            category
            for category in categories
            if lowered in category["name"].lower() or lowered in category["slug"].lower()
        ]
    context = _build_context(
        request,
        {
            "page": "categories",
            "admin": admin,
            "categories": categories,
            "search_term": trimmed,
            "page_message": message,
            "page_message_kind": message_kind,
        },
    )
    return templates.TemplateResponse("admin/categories.html", context)


@router.get("", include_in_schema=False)
def admin_root(request: Request, db: Session = Depends(get_db)):
    if _get_admin_from_session(request, db):
        return RedirectResponse(url="/admin/dashboard", status_code=status.HTTP_303_SEE_OTHER)
    return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/", response_class=HTMLResponse)
def admin_routing(request: Request, db: Session = Depends(get_db)):
    if _get_admin_from_session(request, db):
        return RedirectResponse(url="/admin/dashboard", status_code=status.HTTP_303_SEE_OTHER)
    return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/login", response_class=HTMLResponse)
def login(request: Request, db: Session = Depends(get_db)):
    if _get_admin_from_session(request, db):
        return RedirectResponse(url="/admin/dashboard", status_code=status.HTTP_303_SEE_OTHER)

    return templates.TemplateResponse(
        "admin/login.html",
        _build_context(request, {"page": "login", "form_error": None}),
    )


@router.post("/login", response_class=HTMLResponse)
def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
):
    validate_csrf_token(request, csrf_token)
    normalized_username = username.strip()
    client_id = _client_identifier(request)

    allowed, retry_after = login_rate_limiter.allow(client_id)
    if not allowed:
        logger.warning(
            "admin_login_rate_limited",
            extra={"username": normalized_username, "client": client_id, "retry_after": retry_after},
        )
        context = _build_context(
            request,
            {
                "page": "login",
                "form_error": "Too many login attempts. Please try again in a few minutes.",
            },
        )
        return templates.TemplateResponse("admin/login.html", context, status_code=status.HTTP_429_TOO_MANY_REQUESTS)

    stmt = select(AdminUser).where(
        AdminUser.user_name == normalized_username,
        AdminUser.is_active.is_(True),
    )
    admin = db.execute(stmt).scalar_one_or_none()

    if not admin or not verify_password(password, admin.password_hash):
        logger.warning(
            "admin_login_failed",
            extra={"username": normalized_username, "client": client_id},
        )
        context = _build_context(
            request,
            {"page": "login", "form_error": "Invalid username or password."},
        )
        return templates.TemplateResponse("admin/login.html", context, status_code=status.HTTP_400_BAD_REQUEST)

    request.session["admin_user_id"] = admin.id
    logger.info(
        "admin_login_success",
        extra={"admin_id": admin.id, "username": admin.user_name, "client": client_id},
    )
    return RedirectResponse(url="/admin/dashboard", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/logout")
def logout(request: Request, csrf_token: str = Form(...)):
    validate_csrf_token(request, csrf_token)
    admin_id = request.session.get("admin_user_id")
    request.session.pop("admin_user_id", None)
    logger.info("admin_logout", extra={"admin_user_id": admin_id})
    response = RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)
    return response


@router.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    live_products = db.scalar(select(func.count(Product.id)).where(Product.is_active.is_(True))) or 0
    inactive_products = db.scalar(select(func.count(Product.id)).where(Product.is_active.is_(False))) or 0
    category_count = db.scalar(select(func.count(Category.id))) or 0

    site_visits = _get_metric(db, "site_visits")
    product_views = _get_metric(db, "site_product_views")
    cart_adds = _get_metric(db, "site_cart_adds")
    rate = lambda numerator, denominator: round(numerator / denominator, 2) if denominator else 0

    top_products = db.execute(
        select(Product.name, Product.view_count, Product.cart_add_count)
        .where(Product.is_active.is_(True))
        .order_by(Product.view_count.desc(), Product.cart_add_count.desc(), Product.created_at.desc())
        .limit(5)
    ).all()
    top_categories = db.execute(
        select(Category.name, Category.slug, Category.view_count, Category.cart_add_count)
        .where(Category.is_active.is_(True))
        .order_by(Category.view_count.desc(), Category.cart_add_count.desc(), Category.name.asc())
        .limit(5)
    ).all()

    recent_rows = db.execute(
        select(
            Product.name,
            Product.updated_at,
            Product.is_active,
            Category.name.label("category_name"),
        )
        .outerjoin(Category, Product.category_id == Category.id)
        .order_by(Product.updated_at.desc())
        .limit(5)
    ).all()
    activities = [
        {
            "name": row.name,
            "category": row.category_name,
            "updated_at": row.updated_at or datetime.utcnow(),
            "status": "Active" if row.is_active else "Inactive",
        }
        for row in recent_rows
    ]

    return templates.TemplateResponse(
        "admin/dashboard.html",
        _build_context(
            request,
            {
                "page": "dashboard",
                "admin": admin,
                "stats": {
                    "live_products": live_products,
                    "inactive_products": inactive_products,
                    "categories": category_count,
                },
                "analytics": {
                    "site_visits": site_visits,
                    "product_views": product_views,
                    "cart_adds": cart_adds,
                    "views_per_visit": rate(product_views, site_visits),
                    "cart_adds_per_visit": rate(cart_adds, site_visits),
                    "top_products": top_products,
                    "top_categories": top_categories,
                },
                "activities": activities,
            },
        ),
    )


@router.get("/products", response_class=HTMLResponse)
def manage_products(
    request: Request,
    q: str = Query("", alias="q"),
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    products_query = (
        select(
            Product.id,
            Product.name,
            Product.sku,
            Product.oem_number,
            Product.is_active,
            Category.name.label("category_name"),
            func.count(ProductImage.id).label("image_count"),
        )
        .outerjoin(Category, Product.category_id == Category.id)
        .outerjoin(ProductImage, ProductImage.product_id == Product.id)
        .group_by(
            Product.id,
            Product.name,
            Product.sku,
            Product.oem_number,
            Product.is_active,
            Category.name,
        )
        .order_by(Product.created_at.desc())
    )
    search_term = q.strip()
    if search_term:
        pattern = f"%{search_term}%"
        products_query = products_query.where(
            or_(
                Product.name.ilike(pattern),
                Product.sku.ilike(pattern),
                Product.oem_number.ilike(pattern),
            )
        )

    rows = db.execute(products_query).all()
    products = [
        {
            "id": row.id,
            "name": row.name,
            "sku": row.sku,
            "oem_number": row.oem_number,
            "is_active": row.is_active,
            "category": row.category_name,
            "image_count": row.image_count,
        }
        for row in rows
    ]

    return templates.TemplateResponse(
        "admin/products.html",
        _build_context(
            request,
            {
                "page": "products",
                "admin": admin,
                "products": products,
                "search_term": search_term,
            },
        ),
    )


def _ensure_category_path(db: Session, path: str) -> Optional[int]:
    trimmed = path.strip()
    if not trimmed:
        return None
    parts = [part.strip() for part in trimmed.split(">") if part.strip()]
    parent_id = None
    for depth, part in enumerate(parts):
        existing = db.scalars(
            select(Category).where(Category.parent_id.is_(parent_id), func.lower(Category.name) == part.lower())
        ).first()
        if existing:
            parent_id = existing.id
            continue
        slug_base = _slugify(part)
        slug = slug_base
        counter = 1
        while db.scalar(select(Category).where(Category.slug == slug)):
            counter += 1
            slug = f"{slug_base}-{counter}"
        category = Category(name=part, slug=slug, parent_id=parent_id, level=depth)
        db.add(category)
        db.flush()
        parent_id = category.id
    return parent_id


def _category_path_string(category: Optional[Category]) -> str:
    if not category:
        return ""
    names: List[str] = []
    current = category
    while current:
        names.append(current.name)
        current = current.parent
    return " > ".join(reversed(names))


@router.get("/products/import", response_class=HTMLResponse)
def import_products(request: Request, message: str = "", db: Session = Depends(get_db)):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    context = _build_context(
        request,
        {
            "page": "products",
            "admin": admin,
            "message": message,
        },
    )
    return templates.TemplateResponse("admin/product_import.html", context)


@router.get("/products/import-template")
def download_import_template(request: Request, db: Session = Depends(get_db)):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["Name", "SKU", "OEM Number", "Category Path", "Summary", "Is Active"])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_import_template.xlsx"},
    )


@router.get("/products/export")
def export_products(request: Request, db: Session = Depends(get_db)):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["Name", "SKU", "OEM Number", "Category Path", "Summary", "Is Active"])

    products = db.scalars(
        select(Product).options().order_by(Product.created_at.desc())
    ).all()
    for product in products:
        category_path = _category_path_string(product.category)
        ws.append(
            [
                product.name,
                product.sku,
                product.oem_number,
                category_path,
                product.summary or "",
                "true" if product.is_active else "false",
            ]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_export.xlsx"},
    )


@router.post("/products/import")
async def import_products_post(
    request: Request,
    file: UploadFile = File(...),
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    validate_csrf_token(request, csrf_token)

    if not file.filename.endswith(".xlsx"):
        return RedirectResponse(url="/admin/products/import?message=Upload a .xlsx file", status_code=303)

    data = await file.read()
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = list(first_row[:6]) if first_row else []
    expected = ["Name", "SKU", "OEM Number", "Category Path", "Summary", "Is Active"]
    if headers != expected:
        return RedirectResponse(url="/admin/products/import?message=Headers do not match template", status_code=303)

    created, updated, skipped = 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, sku, oem_number, category_path, summary, is_active = row
        if not (name and sku and oem_number):
            skipped += 1
            continue
        category_id = _ensure_category_path(db, category_path or "") if category_path else None
        product = db.scalar(select(Product).where(Product.sku == sku))
        if product:
            product.name = name
            product.oem_number = oem_number
            product.summary = summary
            product.category_id = category_id
            product.is_active = str(is_active).lower() == "true"
            updated += 1
        else:
            product = Product(
                name=name,
                sku=sku,
                oem_number=oem_number,
                summary=summary,
                category_id=category_id,
                is_active=str(is_active).lower() == "true",
            )
            db.add(product)
            created += 1

    db.commit()
    msg = quote(f"Import complete. Created: {created}, Updated: {updated}, Skipped: {skipped}")
    return RedirectResponse(url=f"/admin/products/import?message={msg}", status_code=303)


@router.get("/leads", response_class=HTMLResponse)
def leads(request: Request, db: Session = Depends(get_db)):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    leads_rows = db.execute(
        select(Lead)
        .order_by(Lead.created_at.desc())
        .limit(100)
    ).scalars().all()

    leads = [
        {
            "id": lead.id,
            "kind": lead.kind,
            "full_name": lead.full_name,
            "email": lead.email,
            "company": lead.company,
            "message": lead.message,
            "created_at": lead.created_at,
        }
        for lead in leads_rows
    ]

    return templates.TemplateResponse(
        "admin/leads.html",
        _build_context(
            request,
            {
                "page": "leads",
                "admin": admin,
                "leads": leads,
            },
        ),
    )


@router.get("/products/new", response_class=HTMLResponse)
def new_product(request: Request, db: Session = Depends(get_db)):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    category_options = _category_parent_options(db, include_predicate=lambda cat: cat.is_active)

    context = _build_context(
        request,
        {
            "page": "products",
            "admin": admin,
            "form_error": None,
            "form_data": {
                "name": "",
                "sku": "",
                "oem_number": "",
                "summary": "",
                "category_id": "",
                "is_active": True,
            },
            "categories": category_options,
            "existing_images": [],
            "form_mode": "create",
            "form_action": "/admin/products/new",
            "submit_label": "Create product",
            "form_title": "Add product",
        },
    )
    return templates.TemplateResponse("admin/product_form.html", context)


@router.post("/products/new", response_class=HTMLResponse)
async def create_product(
    request: Request,
    name: str = Form(...),
    sku: str = Form(...),
    oem_number: str = Form(...),
    summary: str = Form(""),
    category_id: str = Form(""),
    new_images: Union[UploadFile, List[UploadFile], None] = File(default=None),
    csrf_token: str = Form(...),
    is_active: bool = Form(False),
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    validate_csrf_token(request, csrf_token)
    name = name.strip()
    sku = sku.strip()
    oem_number = oem_number.strip()
    summary = summary.strip()
    category_value = category_id.strip() or None

    category_options = _category_parent_options(db, include_predicate=lambda cat: cat.is_active)
    form_data = {
        "name": name,
        "sku": sku,
        "oem_number": oem_number,
        "summary": summary,
        "category_id": category_value or "",
        "is_active": is_active,
    }

    def _render_error(message: str):
        context = _build_context(
            request,
            {
                "page": "products",
                "admin": admin,
                "form_error": message,
                "form_data": form_data,
                "categories": category_options,
                "existing_images": [],
                "form_mode": "create",
                "form_action": "/admin/products/new",
                "submit_label": "Create product",
                "form_title": "Add product",
            },
        )
        return templates.TemplateResponse("admin/product_form.html", context, status_code=status.HTTP_400_BAD_REQUEST)

    if not name or not sku or not oem_number:
        return _render_error("Name, SKU, and OEM number are required.")

    existing = db.scalars(
        select(Product).where(or_(Product.sku == sku, Product.oem_number == oem_number))
    ).first()
    if existing:
        return _render_error("A product with this SKU or OEM number already exists.")

    category_obj = None
    if category_value:
        try:
            category_pk = int(category_value)
        except ValueError:
            return _render_error("Invalid category selection.")

        category_obj = db.get(Category, category_pk)
        if not category_obj:
            return _render_error("Selected category does not exist.")

    product = Product(
        name=name,
        sku=sku,
        oem_number=oem_number,
        summary=summary,
        is_active=is_active,
    )
    if category_obj:
        product.category = category_obj

    upload_files = _coerce_uploads(new_images)

    try:
        uploads = await _process_new_uploads(upload_files, default_alt=name, starting_order=0)
    except ImageUploadError as exc:
        return _render_error(str(exc))
    for image in uploads:
        product.images.append(image)

    db.add(product)
    db.commit()
    logger.info(
        "product_created",
        extra={"product_id": product.id, "admin_id": admin.id},
    )

    return RedirectResponse(url="/admin/products", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/products/{product_id}/edit", response_class=HTMLResponse)
def edit_product(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    product = db.get(Product, product_id)
    if not product:
        return RedirectResponse(url="/admin/products", status_code=status.HTTP_303_SEE_OTHER)

    category_options = _category_parent_options(
        db, include_predicate=lambda cat: cat.is_active or cat.id == (product.category_id or -1)
    )
    sorted_images = sorted(product.images, key=lambda img: img.sort_order)
    images_context = [
        {
            "id": image.id,
            "url": image.image_url,
            "alt_text": image.alt_text or "",
            "sort_order": image.sort_order,
        }
        for image in sorted_images
    ]

    form_data = {
        "name": product.name,
        "sku": product.sku,
        "oem_number": product.oem_number,
        "summary": product.summary or "",
        "category_id": str(product.category_id) if product.category_id else "",
        "is_active": product.is_active,
    }

    context = _build_context(
        request,
        {
            "page": "products",
            "admin": admin,
            "form_error": None,
            "form_data": form_data,
            "categories": category_options,
            "existing_images": images_context,
            "form_mode": "edit",
            "form_action": f"/admin/products/{product_id}/edit",
            "submit_label": "Update product",
            "form_title": f"Edit {product.name}",
            "product_id": product_id,
        },
    )
    return templates.TemplateResponse("admin/product_form.html", context)


@router.post("/products/{product_id}/edit", response_class=HTMLResponse)
async def update_product(
    request: Request,
    product_id: int,
    name: str = Form(...),
    sku: str = Form(...),
    oem_number: str = Form(...),
    summary: str = Form(""),
    category_id: str = Form(""),
    new_images: Union[UploadFile, List[UploadFile], None] = File(default=None),
    csrf_token: str = Form(...),
    is_active: bool = Form(False),
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    product = db.get(Product, product_id)
    if not product:
        return RedirectResponse(url="/admin/products", status_code=status.HTTP_303_SEE_OTHER)

    validate_csrf_token(request, csrf_token)
    name = name.strip()
    sku = sku.strip()
    oem_number = oem_number.strip()
    summary = summary.strip()
    category_value = category_id.strip() or None

    category_options = _category_parent_options(
        db, include_predicate=lambda cat: cat.is_active or cat.id == (product.category_id or -1)
    )
    form_data = {
        "name": name,
        "sku": sku,
        "oem_number": oem_number,
        "summary": summary,
        "category_id": category_value or "",
        "is_active": is_active,
    }

    form_payload = await request.form()

    def _existing_images_context() -> List[Dict[str, Any]]:
        serialized: List[Dict[str, Any]] = []
        for image in sorted(product.images, key=lambda img: img.sort_order):
            serialized.append(
                {
                    "id": image.id,
                    "url": image.image_url,
                    "alt_text": form_payload.get(f"existing_image_alt_{image.id}", image.alt_text or ""),
                    "sort_order": form_payload.get(
                        f"existing_image_order_{image.id}", str(image.sort_order)
                    ),
                    "pending_delete": bool(form_payload.get(f"existing_image_delete_{image.id}")),
                }
            )
        return serialized

    def _render_error(message: str):
        context = _build_context(
            request,
            {
                "page": "products",
                "admin": admin,
                "form_error": message,
                "form_data": form_data,
                "categories": category_options,
                "existing_images": _existing_images_context(),
                "form_mode": "edit",
                "form_action": f"/admin/products/{product_id}/edit",
                "submit_label": "Update product",
                "form_title": f"Edit {product.name}",
                "product_id": product_id,
            },
        )
        return templates.TemplateResponse("admin/product_form.html", context, status_code=status.HTTP_400_BAD_REQUEST)

    if not name or not sku or not oem_number:
        return _render_error("Name, SKU, and OEM number are required.")

    existing = db.scalars(
        select(Product).where(
            or_(Product.sku == sku, Product.oem_number == oem_number),
            Product.id != product.id,
        )
    ).first()
    if existing:
        return _render_error("A product with this SKU or OEM number already exists.")

    category_obj = None
    if category_value:
        try:
            category_pk = int(category_value)
        except ValueError:
            return _render_error("Invalid category selection.")

        category_obj = db.get(Category, category_pk)
        if not category_obj:
            return _render_error("Selected category does not exist.")

    image_updates: List[Tuple[ProductImage, str, int]] = []
    images_to_delete: List[ProductImage] = []
    for image in sorted(product.images, key=lambda img: img.sort_order):
        delete_flag = form_payload.get(f"existing_image_delete_{image.id}")
        alt_value = form_payload.get(f"existing_image_alt_{image.id}", "").strip()
        order_raw = form_payload.get(f"existing_image_order_{image.id}", str(image.sort_order))
        try:
            order_value = int(order_raw)
        except ValueError:
            return _render_error("Image order must be an integer.")

        if delete_flag:
            images_to_delete.append(image)
        else:
            image_updates.append((image, alt_value or name, order_value))

    remaining_orders = [order for (_, _, order) in image_updates]
    max_order = max(remaining_orders) if remaining_orders else -1

    upload_files = _coerce_uploads(new_images)

    try:
        uploads = await _process_new_uploads(
            upload_files,
            default_alt=name,
            starting_order=max_order + 1,
        )
    except ImageUploadError as exc:
        return _render_error(str(exc))

    for image, alt_value, order_value in image_updates:
        image.alt_text = alt_value
        image.sort_order = order_value

    for image in images_to_delete:
        _remove_image_file(image.image_url)
        product.images.remove(image)
        db.delete(image)

    for image in uploads:
        product.images.append(image)

    product.name = name
    product.sku = sku
    product.oem_number = oem_number
    product.summary = summary
    product.category = category_obj
    product.is_active = is_active

    db.add(product)
    db.commit()
    logger.info(
        "product_updated",
        extra={"product_id": product.id, "admin_id": admin.id},
    )

    return RedirectResponse(url="/admin/products", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/categories", response_class=HTMLResponse)
def manage_categories(
    request: Request,
    q: str = Query("", alias="q"),
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    return _render_categories_page(request, admin, db, search_term=q)


@router.get("/categories/new", response_class=HTMLResponse)
def new_category(request: Request, db: Session = Depends(get_db)):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    context = _build_context(
        request,
        {
            "page": "categories",
            "admin": admin,
            "form_error": None,
            "form_data": {"name": "", "slug": "", "description": "", "level": 0, "order": "", "parent_id": "", "is_active": True},
            "parent_options": _category_parent_options(db),
            "form_mode": "create",
            "form_action": "/admin/categories/new",
            "submit_label": "Create category",
            "form_title": "Add category",
            "parent_display": None,
        },
    )
    return templates.TemplateResponse("admin/category_form.html", context)


@router.post("/categories/new", response_class=HTMLResponse)
def create_category(
    request: Request,
    name: str = Form(...),
    slug: str = Form(""),
    description: str = Form(""),
    order: str = Form(""),
    parent_id: str = Form(""),
    csrf_token: str = Form(...),
    is_active: bool = Form(False),
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    validate_csrf_token(request, csrf_token)
    name = name.strip()
    provided_slug = slug.strip()
    description = description.strip()
    order_raw = order.strip()
    parent_raw = parent_id.strip()
    parent_options = _category_parent_options(db)

    form_data = {
        "name": name,
        "slug": provided_slug,
        "description": description,
        "order": order_raw,
        "parent_id": parent_raw,
        "level": "0",
        "is_active": is_active,
    }

    def _render_error(message: str):
        context = _build_context(
            request,
            {
                "page": "categories",
                "admin": admin,
                "form_error": message,
                "form_data": form_data,
                "parent_options": parent_options,
                "form_mode": "create",
                "form_action": "/admin/categories/new",
                "submit_label": "Create category",
                "form_title": "Add category",
                "parent_display": None,
            },
        )
        return templates.TemplateResponse("admin/category_form.html", context, status_code=status.HTTP_400_BAD_REQUEST)


    if not name:
        return _render_error("Category name is required.")

    final_slug = provided_slug or _slugify(name)
    if not final_slug:
        return _render_error("Unable to generate a valid slug. Please provide one manually.")

    existing = db.scalars(select(Category).where(Category.slug == final_slug)).first()
    if existing:
        return _render_error("A category with this slug already exists.")

    parent_obj = None
    if parent_raw:
        try:
            parent_pk = int(parent_raw)
        except ValueError:
            return _render_error("Invalid parent category selection.")

        parent_obj = db.get(Category, parent_pk)
        if not parent_obj:
            return _render_error("Selected parent category does not exist.")

    effective_level = (parent_obj.level + 1) if parent_obj else 0
    form_data["level"] = str(effective_level)
    form_data["parent_id"] = str(parent_obj.id) if parent_obj else ""

    if order_raw:
        try:
            order_value = int(order_raw)
        except ValueError:
            return _render_error("Order must be an integer.")
        form_data["order"] = str(order_value)
    else:
        if parent_obj:
            sibling_filter = Category.parent_id == parent_obj.id
        else:
            sibling_filter = Category.parent_id.is_(None)
        max_order = db.execute(select(func.coalesce(func.max(Category.order), -1)).where(sibling_filter)).scalar_one()
        order_value = max_order + 1
        form_data["order"] = str(order_value)

    category = Category(
        name=name,
        slug=final_slug,
        description=description,
        is_active=is_active,
        order=order_value,
        level=effective_level,
        parent=parent_obj,
    )
    db.add(category)
    db.commit()
    logger.info(
        "category_created",
        extra={"category_id": category.id, "admin_id": admin.id},
    )

    return RedirectResponse(url="/admin/categories", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/categories/{category_id}/edit", response_class=HTMLResponse)
def edit_category(
    request: Request,
    category_id: int,
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    category = db.get(Category, category_id)
    if not category:
        return _render_categories_page(request, admin, db, message="Category not found.", message_kind="error")

    parent = category.parent
    parent_label = parent.name if parent else "No parent (top-level)"
    form_data = {
        "name": category.name,
        "slug": category.slug,
        "description": category.description or "",
        "order": str(category.order),
        "parent_id": str(parent.id) if parent else "",
        "level": str(category.level),
        "is_active": category.is_active,
    }

    context = _build_context(
        request,
        {
            "page": "categories",
            "admin": admin,
            "form_error": None,
            "form_data": form_data,
            "parent_options": None,
            "form_mode": "edit",
            "form_action": f"/admin/categories/{category_id}/edit",
            "submit_label": "Update category",
            "form_title": f"Edit {category.name}",
            "parent_display": parent_label,
            "category_id": category_id,
        },
    )
    return templates.TemplateResponse("admin/category_form.html", context)


@router.post("/categories/{category_id}/edit", response_class=HTMLResponse)
def update_category(
    request: Request,
    category_id: int,
    name: str = Form(...),
    slug: str = Form(""),
    description: str = Form(""),
    order: str = Form(""),
    parent_id: str = Form(""),
    csrf_token: str = Form(...),
    is_active: bool = Form(False),
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    category = db.get(Category, category_id)
    if not category:
        return _render_categories_page(request, admin, db, message="Category not found.", message_kind="error")

    name = name.strip()
    provided_slug = slug.strip()
    description = description.strip()
    order_raw = order.strip()

    parent = category.parent
    parent_label = parent.name if parent else "No parent (top-level)"

    form_data = {
        "name": name,
        "slug": provided_slug,
        "description": description,
        "order": order_raw or str(category.order),
        "parent_id": str(parent.id) if parent else "",
        "level": str(category.level),
        "is_active": is_active,
    }

    validate_csrf_token(request, csrf_token)

    def _render_error(message: str):
        context = _build_context(
            request,
            {
                "page": "categories",
                "admin": admin,
                "form_error": message,
                "form_data": form_data,
                "parent_options": None,
                "form_mode": "edit",
                "form_action": f"/admin/categories/{category_id}/edit",
                "submit_label": "Update category",
                "form_title": f"Edit {category.name}",
                "parent_display": parent_label,
                "category_id": category_id,
            },
        )
        return templates.TemplateResponse("admin/category_form.html", context, status_code=status.HTTP_400_BAD_REQUEST)

    if not name:
        return _render_error("Category name is required.")

    final_slug = provided_slug or _slugify(name)
    if not final_slug:
        return _render_error("Unable to generate a valid slug. Please provide one manually.")

    existing = db.scalars(
        select(Category).where(Category.slug == final_slug, Category.id != category.id)
    ).first()
    if existing:
        return _render_error("A category with this slug already exists.")

    if order_raw:
        try:
            order_value = int(order_raw)
        except ValueError:
            return _render_error("Order must be an integer.")
    else:
        order_value = category.order

    category.name = name
    category.slug = final_slug
    category.description = description
    category.order = order_value
    category.is_active = is_active

    db.add(category)
    db.commit()
    logger.info(
        "category_updated",
        extra={"category_id": category.id, "admin_id": admin.id},
    )

    return RedirectResponse(url="/admin/categories", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/categories/{category_id}/delete", response_class=HTMLResponse)
def delete_category(
    request: Request,
    category_id: int,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    validate_csrf_token(request, csrf_token)
    category = db.get(Category, category_id)
    if not category:
        return _render_categories_page(request, admin, db, message="Category not found.", message_kind="error")

    to_visit = [category]
    collected_ids: List[int] = []
    while to_visit:
        current = to_visit.pop()
        collected_ids.append(current.id)
        to_visit.extend(list(current.children))

    product_count = db.execute(
        select(func.count(Product.id)).where(Product.category_id.in_(collected_ids))
    ).scalar_one()
    if product_count > 0:
        return _render_categories_page(
            request,
            admin,
            db,
            message="Cannot delete a category that still has products assigned. Reassign products first.",
            message_kind="error",
        )

    db.delete(category)
    db.commit()
    logger.info(
        "category_deleted",
        extra={"category_id": category_id, "admin_id": admin.id},
    )

    return RedirectResponse(url="/admin/categories", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/products/{product_id}/delete", response_class=HTMLResponse)
def delete_product(
    request: Request,
    product_id: int,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
):
    admin = _ensure_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_303_SEE_OTHER)

    validate_csrf_token(request, csrf_token)
    prod = db.get(Product, product_id)
    if not prod:
        return RedirectResponse(url="/admin/products", status_code=status.HTTP_303_SEE_OTHER)

    for image in list(prod.images):
        _remove_image_file(image.image_url)

    db.delete(prod)
    db.commit()
    logger.info(
        "product_deleted",
        extra={"product_id": product_id, "admin_id": admin.id},
    )

    return RedirectResponse(url="/admin/products", status_code=status.HTTP_303_SEE_OTHER)
